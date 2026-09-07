from flask import Flask, redirect, render_template, request, url_for, session
import mysql.connector
import time
import os
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import send_file
import bcrypt
from ollama import chat, embed
import json
import math
import base64
import uuid



app =Flask(__name__)
app.secret_key = 'actbayan_secret_key_2026'

def connect_db():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="acttest"
    )


def hash_password_bcrypt(password: str) -> str:
    if password is None:
        password = ""
    pw_bytes = password.encode('utf-8')
    hashed = bcrypt.hashpw(pw_bytes, bcrypt.gensalt())
    return hashed.decode('utf-8')


def check_password_bcrypt(stored_hash, password: str) -> bool:
    if stored_hash is None or password is None:
        return False

    # stored_hash may be stored as TEXT/VARCHAR (string), or bytes (legacy)
    if isinstance(stored_hash, (bytes, bytearray)):
        hash_bytes = bytes(stored_hash)
    else:
        hash_bytes = str(stored_hash).encode('utf-8')

    return bcrypt.checkpw(password.encode('utf-8'), hash_bytes)


# ===== OLLAMA DUPLICATE DETECTION HELPERS =====
def get_embedding(text):
    """Generate embedding vector for text using Ollama nomic-embed-text"""
    try:
        response = embed(
            model='nomic-embed-text',
            input=text
        )
        # Response format: {"embeddings": [[0.1, 0.2, ...]]}
        if response and 'embeddings' in response:
            return response['embeddings'][0]
        return None
    except Exception as e:
        print(f"[Ollama] Embedding error: {e}")
        return None

def cosine_similarity(vec_a, vec_b):
    """Compute cosine similarity between two vectors"""
    if not vec_a or not vec_b or len(vec_a) != len(vec_b):
        return 0.0
    dot = sum(a * b for a, b in zip(vec_a, vec_b))
    norm_a = math.sqrt(sum(a * a for a in vec_a))
    norm_b = math.sqrt(sum(b * b for b in vec_b))
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return dot / (norm_a * norm_b)

def haversine_distance(lat1, lng1, lat2, lng2):
    """Compute distance in meters between two GPS coordinates"""
    R = 6371000  # Earth radius in meters
    phi1 = math.radians(float(lat1))
    phi2 = math.radians(float(lat2))
    delta_phi = math.radians(float(lat2) - float(lat1))
    delta_lambda = math.radians(float(lng2) - float(lng1))
    a = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

def text_similarity(text1, text2):
    """Compare two texts using Ollama embeddings and cosine similarity"""
    if not text1 or not text2:
        return 0.0
    emb1 = get_embedding(text1)
    emb2 = get_embedding(text2)
    if not emb1 or not emb2:
        return 0.0
    return cosine_similarity(emb1, emb2)

def coordinate_proximity(lat1, lng1, lat2, lng2, threshold_meters=100):
    """Check if two coordinates are within threshold meters of each other"""
    try:
        lat1_f = float(lat1)
        lng1_f = float(lng1)
        lat2_f = float(lat2)
        lng2_f = float(lng2)
        if lat1_f == 0 and lng1_f == 0:
            return False, float('inf')
        if lat2_f == 0 and lng2_f == 0:
            return False, float('inf')
        distance = haversine_distance(lat1_f, lng1_f, lat2_f, lng2_f)
        return distance <= threshold_meters, distance
    except (ValueError, TypeError):
        return False, float('inf')


# ===== API: Check for duplicate concerns =====
@app.route('/api/check_duplicate', methods=['POST'])
def api_check_duplicate():
    """Check if a concern with similar title/description and nearby coordinates already exists"""
    if 'user_id' not in session:
        return {'success': False, 'message': 'Not logged in'}, 401
    
    data = request.get_json()
    if not data:
        return {'success': False, 'message': 'No data provided'}, 400
    
    title = (data.get('title') or '').strip()
    description = (data.get('description') or '').strip()
    geox = data.get('geox')
    geoy = data.get('geoy')
    
    if not title and not description:
        return {'success': False, 'message': 'No content to compare'}, 400
    
    # Generate embedding for the new concern text
    combined_text = f"{title}. {description}" if title and description else (title or description)
    
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    
    # Fetch all existing reports with their locations
    cursor.execute("""
        SELECT r.report_id, r.public_key, r.title, r.description, r.category, r.status, l.x, l.y, l.locname
        FROM reports r
        JOIN location l ON r.location_id = l.location_id
        WHERE l.x IS NOT NULL AND l.y IS NOT NULL AND r.public_key IS NOT NULL
        ORDER BY r.created_at DESC
        LIMIT 100
    """)
    existing_reports = cursor.fetchall()
    con.close()
    
    duplicates = []
    
    for report in existing_reports:
        old_title = report.get('title') or ''
        old_desc = report.get('description') or ''
        old_text = f"{old_title}. {old_desc}" if old_title and old_desc else (old_title or old_desc)
        
        # Calculate text similarity
        sim = text_similarity(combined_text, old_text)
        
        # Calculate coordinate proximity
        near, distance = coordinate_proximity(
            geox, geoy,
            report.get('x'), report.get('y')
        )
        
        # Consider duplicate if EITHER condition is strong:
        # - sim >= 0.80 (high text similarity regardless of location)
        # - sim >= 0.70 AND near (moderate text + nearby)
        # - sim >= 0.90 (very high text similarity)
        if sim >= 0.80 or (sim >= 0.70 and near) or sim >= 0.90:
            duplicates.append({
                'report_id': report['report_id'],
                'public_key': report.get('public_key', ''),
                'title': old_title,
                'category': report.get('category', ''),
                'location': report.get('locname', ''),
                'similarity': sim,  # Return as float 0-1
                'distance_m': round(distance, 1) if near else None,
                'status': report.get('status', 'Pending')
            })
    
    # Limit to top 5 most similar
    duplicates.sort(key=lambda d: d['similarity'], reverse=True)
    duplicates = duplicates[:5]
    
    return {
        'success': True,
        'duplicates': duplicates,
        'has_duplicates': len(duplicates) > 0
    }


    
@app.route('/admin', methods=['GET', 'POST'])
def admin(): 
   if 'user_id' not in session or session.get('user_role') != "admin":
       return redirect(url_for("login"))
   
   return redirect(url_for('activities'))


@app.route('/users', methods=['GET', 'POST'])
def users():

    if 'user_id' not in session or session.get('user_role') != 'admin':
        return redirect(url_for('login'))

    con = connect_db()
    cursor = con.cursor(dictionary=True)

    # One row per account/person.
    # Representative credential selection:
    # - Prefer Gmail credential (u.emorph LIKE '%gmail.com') when available
    # - Otherwise fall back to any credential for that account
    #
    # We still show: fullname, profile photo, email, phone number, and current role.
    # Actions must be executed using account_id (not user_id) so the update affects all
    # usercreds under that account/contact.
    cursor.execute(
        """
        SELECT
            a.account_id,
            CONCAT(a.first_name, ' ', a.last_name) AS fullname,
            a.profile_photo,
            c.email,
            c.phone_number,
            r.role,
            -- Representative user_id for display/debug (UI should not use it as identifier)
            rep.user_id AS representative_user_id
        FROM accounts a
        JOIN contacts c ON c.account_id = a.account_id
        JOIN usercreds u_rep_base ON u_rep_base.contact_id = c.contact_id
        -- Pick exactly one credential row per account: prefer Gmail by sorting
        JOIN (
            SELECT
                u.user_id,
                u.contact_id,
                ROW_NUMBER() OVER (
                    PARTITION BY ct.account_id
                    ORDER BY CASE WHEN u.emorph LIKE '%gmail.com' THEN 0 ELSE 1 END,
                             u.user_id ASC
                ) AS rn
            FROM usercreds u
            JOIN contacts ct ON ct.contact_id = u.contact_id
        ) rep ON rep.user_id = u_rep_base.user_id AND rep.rn = 1
        JOIN user_role ur ON ur.user_id = rep.user_id
        JOIN role r ON r.role_id = ur.role_id
        ORDER BY a.account_id ASC
        """
    )
    rows = cursor.fetchall()

    if request.method == 'POST':
        con = connect_db()
        cursor = con.cursor()
        acttype = request.form.get('act')
        account_id = request.form.get('account_id')

        # Update all credentials belonging to the selected account/contact.
        # This ensures a single click affects both resident + Gmail user_ids.
        lock_val = None
        if acttype in ("ban", "restrict", "remove"):
            lock_val = 'Y'
        elif acttype == "unban":
            lock_val = 'N'

        if account_id and lock_val is not None:
            update_sql = """
                UPDATE credentials c
                JOIN usercreds u ON u.user_id = c.user_id
                JOIN contacts ct ON ct.contact_id = u.contact_id
                SET c.is_locked = %s
                WHERE ct.account_id = %s
            """
            cursor.execute(update_sql, (lock_val, account_id))
            con.commit()

        con.close()
        return redirect(url_for("users"))

    con.close()
    return render_template('users.html', users=rows, active_page='users')
    

    

@app.route('/activities')
def activities():
    if 'user_id' not in session or session.get('user_role') != 'admin':
        return redirect(url_for('login'))

    # Filters from query string
    q = (request.args.get('q') or '').strip()
    action_type = (request.args.get('action_type') or '').strip().lower()

    from_date = (request.args.get('from_date') or '').strip()
    to_date = (request.args.get('to_date') or '').strip()

    con = connect_db()
    cursor = con.cursor(dictionary=True)

    activities = []

    # Common WHERE fragments
    def build_where(alias_dt_field: str, alias_title: str, alias_fullname: str, alias_type_literal: str):
        clauses = []
        params = []

        # action_type filter (works as an overall filter on the activity type)
        if action_type:
            clauses.append(f"{alias_type_literal} = %s")
            params.append(action_type)

        if q:
            # Match either title OR fullname using a single search box.
            clauses.append(f"(LOWER({alias_title}) LIKE %s OR LOWER({alias_fullname}) LIKE %s)")
            needle = f"%{q.lower()}%"
            params.append(needle)
            params.append(needle)


        # Date range: compare against the activity datetime field
        if from_date:
            clauses.append(f"DATE({alias_dt_field}) >= %s")
            params.append(from_date)

        if to_date:
            clauses.append(f"DATE({alias_dt_field}) <= %s")
            params.append(to_date)

        if not clauses:
            return "", []

        return "WHERE " + " AND ".join(clauses), params

    # 1) Reports created (type='report')
    where_report, params_report = build_where('r.created_at', 'r.title', "CONCAT(a.first_name, ' ', a.last_name)", "'report'")
    cursor.execute(
        f"""
        SELECT
            r.created_at AS dt,
            'report' AS type,
            r.title AS title,
            CONCAT(a.first_name, ' ', a.last_name) AS fullname,
            NULL AS text
        FROM reports r
        JOIN accounts a ON r.account_id = a.account_id
        {where_report}
        """,
        tuple(params_report)
    )
    activities.extend(cursor.fetchall())

    # 2) Report updates (type='update')
    where_update, params_update = build_where('rp.submitted_at', 'r.title', "CONCAT(a.first_name, ' ', a.last_name)", "'update'")
    cursor.execute(
        f"""
        SELECT
            rp.submitted_at AS dt,
            'update' AS type,
            r.title AS title,
            CONCAT(a.first_name, ' ', a.last_name) AS fullname,
            rp.description AS text
        FROM report_update rp
        JOIN reports r ON rp.report_id = r.report_id
        JOIN accounts a ON rp.account_id = a.account_id
        {where_update}
        """,
        tuple(params_update)
    )
    activities.extend(cursor.fetchall())

    # 3) Feedback uploads (type='feedback')
    where_feedback, params_feedback = build_where('f.uploaded_at', 'r.title', "CONCAT(a.first_name, ' ', a.last_name)", "'feedback'")
    cursor.execute(
        f"""
        SELECT
            f.uploaded_at AS dt,
            'feedback' AS type,
            r.title AS title,
            CONCAT(a.first_name, ' ', a.last_name) AS fullname,
            f.feedback AS text
        FROM feedbacks f
        JOIN reports r ON f.report_id = r.report_id
        JOIN accounts a ON f.account_id = a.account_id
        {where_feedback}
        """,
        tuple(params_feedback)
    )
    activities.extend(cursor.fetchall())

    con.close()

    # Sort descending by datetime
    activities.sort(key=lambda x: x.get('dt') or datetime.min, reverse=True)

    # Format dt for template
    for a in activities:
        dt = a.get('dt')
        a['dt'] = dt.strftime('%Y-%m-%d %H:%M:%S') if dt else ''

    return render_template('activities.html', activities=activities, active_page='activities')

# --- EXPORT ROUTES ---
@app.route('/api/export/users')
def export_users():
    if 'user_id' not in session or session.get('user_role') != 'admin':
        return redirect(url_for('login'))

    con = connect_db()
    cursor = con.cursor(dictionary=True)
    cursor.execute("""
        SELECT a.account_id, CONCAT(a.first_name, " ", a.last_name) AS fullname, 
               c.email, c.phone_number, r.role 
        FROM accounts a
        JOIN contacts c ON c.account_id = a.account_id 
        JOIN usercreds u ON c.contact_id = u.contact_id
        JOIN user_role ur ON u.user_id = ur.user_id
        JOIN role r ON ur.role_id = r.role_id
        ORDER BY a.account_id ASC
    """)
    users = cursor.fetchall()
    con.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Accounts"

    # Headers
    headers = ["Account ID", "Full Name", "Email Address", "Phone Number", "Role"]
    ws.append(headers)

    # Styling
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data
    for i, u in enumerate(users, start=2):
        ws.cell(row=i, column=1, value=u['account_id'])
        ws.cell(row=i, column=2, value=u['fullname'])
        ws.cell(row=i, column=3, value=u['email'])
        ws.cell(row=i, column=4, value=u['phone_number'])
        ws.cell(row=i, column=5, value=u['role'])
        
        for col in range(1, 6):
            cell = ws.cell(row=i, column=col)
            cell.border = thin_border
            if col in [1, 5]:
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"ActBayan_Users_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/api/export/activities')
def export_activities():
    if 'user_id' not in session or session.get('user_role') != 'admin':
        return redirect(url_for('login'))

    # Filters (same as activities route)
    q = (request.args.get('q') or '').strip()
    action_type = (request.args.get('action_type') or '').strip().lower()
    from_date = (request.args.get('from_date') or '').strip()
    to_date = (request.args.get('to_date') or '').strip()

    con = connect_db()
    cursor = con.cursor(dictionary=True)
    activities = []

    def build_where(alias_dt_field, alias_title, alias_fullname, alias_type_literal):
        clauses = []
        params = []
        if action_type:
            clauses.append(f"{alias_type_literal} = %s")
            params.append(action_type)
        if q:
            clauses.append(f"(LOWER({alias_title}) LIKE %s OR LOWER({alias_fullname}) LIKE %s)")
            needle = f"%{q.lower()}%"
            params.extend([needle, needle])
        if from_date:
            clauses.append(f"DATE({alias_dt_field}) >= %s")
            params.append(from_date)
        if to_date:
            clauses.append(f"DATE({alias_dt_field}) <= %s")
            params.append(to_date)
        
        where_sql = " AND ".join(clauses)
        if where_sql:
            where_sql = "WHERE " + where_sql
        return where_sql, params

    # Reports
    w_sql, w_params = build_where('r.created_at', 'r.title', 'CONCAT(a.first_name," ",a.last_name)', "'report'")
    q1 = f"""
        SELECT r.created_at AS dt, 'Report' AS type, r.title, CONCAT(a.first_name,' ',a.last_name) AS fullname, '' AS text
        FROM reports r
        LEFT JOIN accounts a ON r.account_id = a.account_id
        {w_sql}
    """
    cursor.execute(q1, tuple(w_params))
    activities.extend(cursor.fetchall())

    # Feedbacks
    w_sql2, w_params2 = build_where('f.uploaded_at', 'r.title', 'CONCAT(a.first_name," ",a.last_name)', "'feedback'")
    q2 = f"""
        SELECT f.uploaded_at AS dt, 'Feedback' AS type, r.title, CONCAT(a.first_name,' ',a.last_name) AS fullname, f.feedback AS text
        FROM feedbacks f
        LEFT JOIN reports r ON f.report_id = r.report_id
        LEFT JOIN accounts a ON f.account_id = a.account_id
        {w_sql2}
    """
    cursor.execute(q2, tuple(w_params2))
    activities.extend(cursor.fetchall())

    # Updates
    w_sql3, w_params3 = build_where('rp.submitted_at', 'r.title', 'CONCAT(a.first_name," ",a.last_name)', "'update'")
    q3 = f"""
        SELECT rp.submitted_at AS dt, 'Update' AS type, r.title, CONCAT(a.first_name,' ',a.last_name) AS fullname, rp.description AS text
        FROM report_update rp
        LEFT JOIN reports r ON rp.report_id = r.report_id
        LEFT JOIN accounts a ON rp.account_id = a.account_id
        {w_sql3}
    """
    cursor.execute(q3, tuple(w_params3))
    activities.extend(cursor.fetchall())

    con.close()

    activities.sort(key=lambda x: x['dt'] if x['dt'] else datetime.min, reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Activities"

    headers = ["Date & Time", "Action Type", "Title/Report", "Full Name", "Description"]
    ws.append(headers)

    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for i, act in enumerate(activities, start=2):
        dt_val = act['dt'].strftime('%Y-%m-%d %H:%M:%S') if act.get('dt') else ''
        ws.cell(row=i, column=1, value=dt_val)
        ws.cell(row=i, column=2, value=act.get('type', ''))
        ws.cell(row=i, column=3, value=act.get('title', ''))
        ws.cell(row=i, column=4, value=act.get('fullname', ''))
        ws.cell(row=i, column=5, value=act.get('text', ''))
        
        for col in range(1, 6):
            cell = ws.cell(row=i, column=col)
            cell.border = thin_border
            if col in [1, 2]:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True if col == 5 else False)

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 60

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"ActBayan_Activities_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/api/comments/<int:report_id>')
def api_comments(report_id):
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    cursor.execute("""
        SELECT
            f.feedback AS comment_text,
            a.first_name,
            a.last_name
        FROM feedbacks f
        LEFT JOIN accounts a ON a.account_id = f.account_id
        WHERE f.report_id = %s
        ORDER BY f.uploaded_at DESC
    """, (report_id,))
    comments = cursor.fetchall()
    con.close()
    
    from flask import jsonify
    return jsonify(comments)

@app.route('/api/report_updates/<int:report_id>')
def api_report_updates(report_id):
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    cursor.execute("""
        SELECT description AS update_text, image_url, submitted_at
        FROM report_update
        WHERE report_id = %s
        ORDER BY submitted_at ASC
    """, (report_id,))
    updates = cursor.fetchall()
    con.close()
    
    # Format dates for frontend
    for u in updates:
        if u['submitted_at']:
            u['created_at_formatted'] = u['submitted_at'].strftime('%B %d, %Y - %I:%M %p')
            # delete datetime object since it cannot be serialized directly by flask jsonify (unless configured)
            del u['submitted_at']
            
    from flask import jsonify
    return jsonify(updates)
    
@app.route('/admin/login_attempts')
def admin_login_attempts():
    if 'user_id' not in session or session.get('user_role') != 'admin':
        return redirect(url_for('login'))
        
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    cursor.execute("SELECT * FROM login_logs ORDER BY attempt_time DESC")
    logs = cursor.fetchall()
    
    cursor.execute("SELECT c.user_id, c.failed_attempts, c.is_locked, c.lock_duration, a.first_name, a.last_name, a.profile_photo FROM credentials c JOIN usercreds u ON c.user_id = u.user_id JOIN contacts ct ON u.contact_id = ct.contact_id JOIN accounts a ON ct.account_id = a.account_id WHERE c.failed_attempts > 0 OR c.is_locked = 'Y'")
    banned_users = cursor.fetchall()
    
    con.close()
    return render_template('admin_login_attempts.html', logs=logs, banned_users=banned_users, active_page='login_attempts')

@app.route('/admin/reports')
def admin_reports():
    if 'user_id' not in session or session.get('user_role') != 'admin':
        return redirect(url_for('login'))
        
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    cursor.execute("""
        SELECT r.*, a.first_name, a.last_name, a.profile_photo AS reporter_photo, l.locname 
        FROM reports r 
        LEFT JOIN accounts a ON r.account_id = a.account_id 
        LEFT JOIN location l ON r.location_id = l.location_id
        ORDER BY r.created_at DESC
    """)
    reports = cursor.fetchall()
    con.close()
    return render_template('admin_reports.html', reports=reports, active_page='reports')

@app.route('/admin/settings', methods=['GET', 'POST'])
def admin_settings():
    if 'user_id' not in session or session.get('user_role') != 'admin':
        return redirect(url_for('login'))
        
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'change_role':
            target_user_id = request.form.get('usercreds_user_id')
            new_role_id = request.form.get('role_id')
            if target_user_id and new_role_id:
                # Need to update or insert?
                cursor.execute("SELECT 1 FROM user_role WHERE user_id = %s", (target_user_id,))
                exists = cursor.fetchone()
                if exists:
                    contactlist = """
                    SELECT contact_id FROM usercreds WHERE user_id = %s
                    """
                    
                    cursor.execute(contactlist, (target_user_id,))
                    c = cursor.fetchone()
                    
                    cursor.execute("SELECT user_id FROM usercreds WHERE contact_id = %s", (c["contact_id"],))
                    us = cursor.fetchall()
                    
                    for users in us:
                        cursor.execute("UPDATE user_role SET role_id = %s WHERE user_id = %s", (new_role_id, users["user_id"]))
                else:
                    cursor.execute("INSERT INTO user_role (user_id, role_id) VALUES (%s, %s)", (target_user_id, new_role_id))
                con.commit()
    
    cursor.execute("""
        SELECT 
    u.user_id,
    a.first_name,
    a.last_name,
    a.profile_photo,
    r.role_id,
    r.role
    FROM usercreds u
    JOIN contacts ct ON u.contact_id = ct.contact_id
    JOIN accounts a ON ct.account_id = a.account_id
    LEFT JOIN user_role ur ON u.user_id = ur.user_id
    LEFT JOIN role r ON ur.role_id = r.role_id
    WHERE u.emorph LIKE '%gmail.com';
    """)
    users = cursor.fetchall()
    
    cursor.execute("SELECT * FROM role")
    roles = cursor.fetchall()
    con.close()
    return render_template('admin_settings.html', users=users, roles=roles, active_page='settings')

@app.route('/')
def home():
    user_data = None
    if 'user_id' in session:
        try:
            con = connect_db()
            cursor = con.cursor(dictionary=True)
            cursor.execute("SELECT first_name, last_name, profile_photo FROM accounts WHERE account_id = %s", (session['user_id'],))
            rows = cursor.fetchall()
            if rows:
                user_data = rows[0]
            con.close()
        except Exception:
            pass
    return render_template('index.html', user=user_data)

@app.route('/lgu', methods=['GET', 'POST'])
def lgu():
    if 'user_id' not in session or session.get('user_role') != 'lgu':
        return redirect(url_for('login'))


    account_id = session.get('user_id')

    # Fetch data for GET (and for re-render if needed)
    sql_reps = """
        SELECT r.*, a.first_name, a.last_name, a.profile_photo
        FROM reports r
        JOIN accounts a ON r.account_id = a.account_id
        ORDER BY r.created_at DESC
    """

    if request.method == 'POST':
        data = request.form
        report_id = data.get('report_id')
        description = data.get('update_description', '').strip()

        # Basic validation
        if not report_id or not description:
            return redirect(url_for('lgu'))

        # Optional image upload
        image_url = None
        update_photo = request.files.get('update_photo')
        if update_photo and update_photo.filename:
            os.makedirs('static/uploads', exist_ok=True)
            filename = f"update_{report_id}_{int(datetime.now().timestamp())}_{update_photo.filename}"
            filepath = f"static/uploads/{filename}"
            update_photo.save(filepath)
            image_url = filepath

        con = connect_db()
        cursor = con.cursor()
        query = """
            INSERT INTO report_update (account_id, report_id, image_url, description, submitted_at)
            VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(query, (
            account_id,
            report_id,
            image_url,
            description,
            datetime.now()
        ))
        
        status = data.get('update_status')
        if status:
            cursor.execute("UPDATE reports SET status = %s WHERE report_id = %s", (status, report_id))
            
        con.commit()
        con.close()
        return redirect(url_for('lgu'))

    # GET
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    cursor.execute(sql_reps)
    reps = cursor.fetchall()

    report_ids = [r.get('report_id') for r in reps if r.get('report_id') is not None]

    # ===== FEEDBACKS / COMMENTS (same query as file_concern) =====
    feedbacks_by_report = {}
    feedback_counts_by_report = {}

    if report_ids:
        placeholders = ','.join(['%s'] * len(report_ids))
        feedback_query = f"""
            SELECT
                f.feedback_id,
                f.report_id,
                f.account_id,
                f.feedback,
                f.uploaded_at,
                a.profile_photo,
                CONCAT(a.first_name, ' ', a.last_name) AS fullname
            FROM feedbacks f
            LEFT JOIN accounts a ON a.account_id = f.account_id
            WHERE f.report_id IN ({placeholders})
            ORDER BY f.feedback_id DESC
        """
        cursor.execute(feedback_query, tuple(report_ids))
        all_feedbacks = cursor.fetchall()
        now = datetime.now()
        for fb in all_feedbacks:
            if fb.get('uploaded_at'):
                diff = now - fb['uploaded_at']
                if diff.days > 7:
                    fb['time_ago'] = fb['uploaded_at'].strftime('%b %d, %Y')
                elif diff.days > 0:
                    fb['time_ago'] = f"{diff.days}d ago"
                else:
                    seconds = diff.seconds
                    if seconds >= 3600:
                        fb['time_ago'] = f"{seconds // 3600}h ago"
                    elif seconds >= 60:
                        fb['time_ago'] = f"{seconds // 60}m ago"
                    else:
                        fb['time_ago'] = "Just now"
            else:
                fb['time_ago'] = ""
            feedbacks_by_report.setdefault(fb.get('report_id'), []).append(fb)

    for rid, fbs in feedbacks_by_report.items():
        feedback_counts_by_report[rid] = len(fbs)

    # ===== UPVOTES (same query as file_concern) =====
    upvote_counts = {}
    user_upvoted = set()

    if report_ids:
        placeholders = ','.join(['%s'] * len(report_ids))
        cursor.execute(
            f"SELECT report_id, COUNT(*) as count FROM upvotes WHERE report_id IN ({placeholders}) GROUP BY report_id",
            tuple(report_ids)
        )
        for row in cursor.fetchall():
            upvote_counts[row['report_id']] = row['count']

        # LGU also tracks their own upvotes
        if account_id:
            cursor.execute(
                f"SELECT report_id FROM upvotes WHERE account_id = %s AND report_id IN ({placeholders})",
                (account_id,) + tuple(report_ids)
            )
            for row in cursor.fetchall():
                user_upvoted.add(row['report_id'])

    # ===== LGU USER PROFILE (for sidebar) =====
    cursor.execute(
        "SELECT first_name, last_name, profile_photo FROM accounts WHERE account_id = %s",
        (account_id,)
    )
    user_data = cursor.fetchone()
    con.close()

    return render_template(
        "lgu.html",
        reps=reps,
        feedbacks_by_report=feedbacks_by_report,
        feedback_counts_by_report=feedback_counts_by_report,
        upvote_counts=upvote_counts,
        user_upvoted=user_upvoted,
        user=user_data
    )



@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    con = connect_db()
    cursor = con.cursor(dictionary=True)

    account_id = session.get('user_id')
    user_name = session.get('user_name')

    sql_recent = """
        SELECT r.*, a.first_name, a.last_name 
        FROM reports r
        JOIN accounts a ON r.account_id = a.account_id
        ORDER BY r.created_at DESC
    """
    cursor.execute(sql_recent)
    all_concerns = cursor.fetchall()

    # Fetch real user data for sidebar
    cursor.execute(
        "SELECT first_name, last_name, profile_photo, account_status FROM accounts WHERE account_id = %s",
        (account_id,)
    )
    user_rows = cursor.fetchall()
    user_data = user_rows[0] if user_rows else None
    
    cursor.execute("SELECT * FROM announcements ORDER BY submitted_at DESC LIMIT 1")
    ann_rows = cursor.fetchall()
    latest_ann = ann_rows[0] if ann_rows else None
    if latest_ann and latest_ann['submitted_at']:
        latest_ann['formatted_date'] = latest_ann['submitted_at'].strftime('%b %d, %Y')
        
    con.close()

    return render_template(
        'dashboard.html',
        recent=all_concerns,
        ngalan=user_name,
        user=user_data,
        latest_ann=latest_ann
    )


@app.route('/login', methods=['GET', 'POST'])
def login():
    con = connect_db()
    cursor = con.cursor(buffered=True)

    if request.method == 'POST':
        data = request.form
        u = data.get('e', '').strip()
        p = data.get('pass', '').strip()
        
        if not u or not p:
            return render_template('login.html', error="Please enter email/phone and password")
        
        # Fetch registration record
        query_reg = "SELECT registration_id, email_address, phone_number, password FROM registration WHERE email_address = %s OR phone_number = %s"
        cursor.execute(query_reg, (u, u))
        reg_record = cursor.fetchone()
        
        if reg_record and check_password_bcrypt(reg_record[3], p):

            reg_id = reg_record[0]
            if reg_id:
                return redirect(url_for('cd', reg_id=reg_id))
        
        credquery = """SELECT u.emorph, p.password FROM credentials AS c
                    JOIN usercreds AS u ON u.user_id = c.user_id
                    JOIN password AS p ON p.pass_id = c.pass_id
                    WHERE u.emorph = %s
                    """
        cursor.execute(credquery, (u,))
        creds = cursor.fetchone()
        if creds and check_password_bcrypt(creds[1], p):

            emorph = creds[0]

            neym = """SELECT c.account_id, a.first_name
              FROM accounts a
              JOIN contacts c ON a.account_id = c.account_id
              WHERE c.email = %s OR c.phone_number = %s
           """
            cursor.execute(neym, (emorph, emorph))
            nem = cursor.fetchone()

            if nem:
                account_id = nem[0]
                pangalan = nem[1]
                
                # SELECT u.user_id, u.emorph, r.role FROM usercreds u JOIN contacts c ON u.contact_id = c.contact_id JOIN user_role ur ON u.user_id = ur.user_id JOIN role r ON ur.role_id = r.role_id WHERE u.emorph = %s;
                
                # Role verification (check using email/phone first -> emorph -> usercreds.user_id)
                rowl = """
                    SELECT u.user_id, u.emorph, r.role
                    FROM usercreds u
                    JOIN contacts c ON u.contact_id = c.contact_id
                    JOIN user_role ur ON u.user_id = ur.user_id
                    JOIN role r ON ur.role_id = r.role_id
                    WHERE u.emorph = %s
                """
                cursor.execute(rowl, (emorph,))
                rul = cursor.fetchone()

                # If user is found in user_role table, check banned status from credentials
                usercreds_user_id = None
                role_value = ""
                if rul:
                    usercreds_user_id = rul[0]
                    if len(rul) > 2 and rul[2] is not None:
                        role_value = str(rul[2]).strip().casefold()

                if usercreds_user_id is not None:
                    ban_check_sql = "SELECT is_locked FROM credentials WHERE user_id = %s"
                    cursor.execute(ban_check_sql, (usercreds_user_id,))
                    ban_row = cursor.fetchone()
                    is_locked = ban_row[0] if ban_row else None
                    if is_locked == 'Y':
                        cursor.execute("INSERT INTO login_logs (emorph, status, attempt_time) VALUES (%s, %s, %s)", (u, 'Failed (Banned)', datetime.now()))
                        con.commit()
                        con.close()
                        return render_template('login.html', error="This account is banned.")

                cursor.execute("INSERT INTO login_logs (emorph, status, attempt_time) VALUES (%s, %s, %s)", (u, 'Success', datetime.now()))
                con.commit()

                session['user_id'] = account_id
                session['user_name'] = pangalan

                if role_value == 'lgu official':
                    # IMPORTANT: `/lgu_announcements` expects session['user_role'] == 'lgu'
                    session['user_role'] = 'lgu'
                    con.close()
                    return redirect(url_for("lgu"))
                elif role_value == "admin":
                    session['user_role'] = 'admin'
                    con.close()
                    return redirect(url_for("admin"))
                
                session['user_role'] = 'resident'
                con.close()
                return redirect(url_for("dashboard"))
            

        cursor.execute("INSERT INTO login_logs (emorph, status, attempt_time) VALUES (%s, %s, %s)", (u, 'Failed', datetime.now()))
        con.commit()
        con.close()
        return render_template('login.html', error="Invalid credentials")
    return render_template('login.html')


@app.route('/file_concern', methods=['GET', 'POST'])
def file_concern():

    if 'user_id' not in session:
        return redirect(url_for('login'))

    con = connect_db()
    cursor = con.cursor(dictionary=True)

    if request.method == 'POST':
        con.autocommit = False
        data = request.form
        fn = request.form.get("fn")
        
        category = data.get('category')
        title = data.get('title')
        # Template sends location text (locname)
        location = data.get('location')
        description = data.get('description')

        # Optional coordinates (template currently only sends location text;
        # if geox/geoy are missing, we try to default to 0 so location_id is not NULL)
        geox = data.get('geox')
        geoy = data.get('geoy')

        if geox is None or str(geox).strip() == '':
            geox = 0
        if geoy is None or str(geoy).strip() == '':
            geoy = 0
        
        cursor.execute("SELECT account_id, account_status FROM accounts WHERE account_id = %s", (session['user_id'],))
        user_rows = cursor.fetchall()
        user_record = user_rows[0] if user_rows else None

        if not user_record:
            return redirect(url_for('login'))
            
        if user_record.get('account_status') != 'Verified':
            con.close()
            return redirect(url_for('file_concern'))
            
        account_id = user_record['account_id']
        
        image_file = request.files.get('concern_photo')
        img_url = None
        if image_file and image_file.filename:
            os.makedirs('static/uploads/concerns', exist_ok=True)
            filename = f"concern_{account_id}_{int(datetime.now().timestamp())}_{image_file.filename}"
            filepath = f"static/uploads/concerns/{filename}"
            image_file.save(filepath)
            img_url = filepath
            
        # For file_concern: insert geolocation into `location` first,
        # then insert report using the resulting `location_id`.
        # Expected form fields: location (locname), geox, geoy.
        insert_location_sql = """
            INSERT INTO location (locname, x, y)
            VALUES (%s, %s, %s)
        """
        get_location_id_sql = "SELECT MAX(location_id) FROM location"  # Get the last inserted location_id for association with report
        
        query = """
            INSERT INTO reports (account_id, category, title, location_id, image_url, description, status, public_key)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """

        # For fn==pc (posting feedback), do NOT insert a new report
        if fn == "pc":
            feedback = (data.get('comment') or '').strip()
            report_id = data.get('report_id')

            if report_id and feedback:
                comque = """
                    INSERT INTO feedbacks (account_id, report_id, feedback, uploaded_at)
                    VALUES (%s, %s, %s, %s)
                """
                cursor.execute(comque, (account_id, report_id, feedback, datetime.now()))

        else:
            # If geolocation coordinates are provided, store them in `location`.
            # - location variable from template is treated as locname
            # - request.form should provide geox and geoy
            # (fallbacks included because some browsers/templates may not send them)
            geox = data.get('geox')
            geoy = data.get('geoy')

            location_id = None
            # Always insert into `location` so we can always attach a non-NULL location_id.
            # If geox/geoy were missing, we already defaulted them to 0 above.
            cursor.execute(insert_location_sql, (location, geox, geoy))
            cursor.execute(get_location_id_sql)
            location_row = cursor.fetchone()
            location_id = None
            if location_row:
                # cursor is dictionary=True, so fetchone() returns a dict like {"MAX(location_id)": 123}
                location_id = location_row.get('MAX(location_id)')

            # Generate UUID public key for this report
            public_key = str(uuid.uuid4())
            cursor.execute(query, (account_id, category, title, location_id, img_url, description, 'Pending', public_key))

        con.commit()
        con.close()
        return redirect(url_for('file_concern'))
        
    fetch_query = """
        SELECT r.*, a.first_name, a.last_name, a.profile_photo
        FROM reports r
        JOIN accounts a ON r.account_id = a.account_id
        ORDER BY r.created_at DESC
    """
    cursor.execute(fetch_query)
    all_concerns = cursor.fetchall()

    # Fetch all updates and group them by report_id for per-card rendering
    cursor.execute("""
        SELECT 
        CONCAT(a.first_name, ' ', a.last_name) AS fullname,
        rp.report_id,
        rp.image_url,
        rp.description,
        rp.submitted_at
        FROM report_update rp
        JOIN accounts a ON a.account_id = rp.account_id
        ORDER BY rp.submitted_at DESC;
    """)
    all_updates = cursor.fetchall()

    # Fetch feedbacks for all reports in one query and group them by report_id
    report_ids = [c.get('report_id') for c in all_concerns if c.get('report_id') is not None]
    feedbacks_by_report = {}
    feedback_counts_by_report = {}

    if report_ids:
        placeholders = ",".join(["%s"] * len(report_ids))
        feedback_query = f"""
            SELECT 
                f.feedback_id,
                f.report_id,
                f.account_id,
                f.feedback,
                f.uploaded_at,
                a.profile_photo,
                CONCAT(a.first_name, ' ', a.last_name) AS fullname
            FROM feedbacks f
            LEFT JOIN accounts a ON a.account_id = f.account_id
            WHERE f.report_id IN ({placeholders})
            ORDER BY f.feedback_id DESC
        """
        cursor.execute(feedback_query, tuple(report_ids))
        all_feedbacks = cursor.fetchall()
        now = datetime.now()
        for fb in all_feedbacks:
            if fb.get('uploaded_at'):
                diff = now - fb['uploaded_at']
                if diff.days > 7:
                    fb['time_ago'] = fb['uploaded_at'].strftime('%b %d, %Y')
                elif diff.days > 0:
                    fb['time_ago'] = f"{diff.days}d ago"
                else:
                    seconds = diff.seconds
                    if seconds >= 3600:
                        hours = seconds // 3600
                        fb['time_ago'] = f"{hours}h ago"
                    elif seconds >= 60:
                        minutes = seconds // 60
                        fb['time_ago'] = f"{minutes}m ago"
                    else:
                        fb['time_ago'] = "Just now"
            else:
                fb['time_ago'] = ""
                
            rid = fb.get('report_id')
            feedbacks_by_report.setdefault(rid, []).append(fb)

    # Build counts for comment badges
    for rid, fbs in feedbacks_by_report.items():
        feedback_counts_by_report[rid] = len(fbs)

    updates_by_report = {}
    for upd in all_updates:
        rid = upd.get('report_id')
        if rid is None:
            continue
        updates_by_report.setdefault(rid, []).append(upd)

    upvote_counts = {}
    user_upvoted = set()

    account_id = session.get('user_id')

    if report_ids:
        # Get upvote counts
        upvote_query = f"SELECT report_id, COUNT(*) as count FROM upvotes WHERE report_id IN ({placeholders}) GROUP BY report_id"
        cursor.execute(upvote_query, tuple(report_ids))
        for row in cursor.fetchall():
            upvote_counts[row['report_id']] = row['count']
            
        # Get user's upvotes
        if account_id:
            user_upvote_query = f"SELECT report_id FROM upvotes WHERE account_id = %s AND report_id IN ({placeholders})"
            cursor.execute(user_upvote_query, (account_id,) + tuple(report_ids))
            for row in cursor.fetchall():
                user_upvoted.add(row['report_id'])

    # Fetch real user data for sidebar
    cursor.execute(
        "SELECT first_name, last_name, profile_photo, account_status FROM accounts WHERE account_id = %s",
        (account_id,)
    )
    user_data = cursor.fetchone()
    con.close()

    return render_template(
        'file_concern.html',
        concerns=all_concerns,
        updates_by_report=updates_by_report,
        feedbacks_by_report=feedbacks_by_report,
        feedback_counts_by_report=feedback_counts_by_report,
        upvote_counts=upvote_counts,
        user_upvoted=user_upvoted,
        user=user_data
    )
# ===== API: Fetch report details by public_key (UUID) =====
@app.route('/api/report/<public_key>')
def api_report_by_key(public_key):
    """Fetch full report details by its UUID public_key for the modal viewer"""
    if 'user_id' not in session:
        from flask import jsonify
        return jsonify({'success': False, 'message': 'Not logged in'}), 401
    
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    cursor.execute("""
        SELECT r.report_id, r.public_key, r.title, r.description, r.category, r.status,
               r.image_url, r.created_at,
               a.first_name, a.last_name, a.profile_photo,
               l.locname, l.x, l.y
        FROM reports r
        JOIN accounts a ON r.account_id = a.account_id
        LEFT JOIN location l ON r.location_id = l.location_id
        WHERE r.public_key = %s
    """, (public_key,))
    report = cursor.fetchone()
    
    if not report:
        con.close()
        from flask import jsonify
        return jsonify({'success': False, 'message': 'Report not found'}), 404
    
    # Sanitize bytes
    for k, v in list(report.items()):
        if isinstance(v, (bytes, bytearray)):
            report[k] = v.decode('utf-8', errors='ignore') if v else ""
        elif isinstance(v, datetime):
            report[k] = v.strftime('%B %d, %Y - %I:%M %p')
    
    # Get upvote count
    cursor.execute("SELECT COUNT(*) as count FROM upvotes WHERE report_id = %s", (report['report_id'],))
    report['upvote_count'] = cursor.fetchone()['count']
    
    # Get comment count
    cursor.execute("SELECT COUNT(*) as count FROM feedbacks WHERE report_id = %s", (report['report_id'],))
    report['comment_count'] = cursor.fetchone()['count']
    
    # Get recent comments
    cursor.execute("""
        SELECT f.feedback, f.uploaded_at,
               CONCAT(a.first_name, ' ', a.last_name) AS fullname,
               a.profile_photo
        FROM feedbacks f
        LEFT JOIN accounts a ON a.account_id = f.account_id
        WHERE f.report_id = %s
        ORDER BY f.uploaded_at DESC
        LIMIT 10
    """, (report['report_id'],))
    comments = cursor.fetchall()
    for c in comments:
        for k, v in list(c.items()):
            if isinstance(v, (bytes, bytearray)):
                c[k] = v.decode('utf-8', errors='ignore') if v else ""
            elif isinstance(v, datetime):
                c[k] = v.strftime('%b %d, %Y • %I:%M %p')
    
    report['comments'] = comments
    
    con.close()
    from flask import jsonify
    return jsonify({'success': True, 'report': report})


@app.route('/api/upvote', methods=['POST'])
def api_upvote():
    if 'user_id' not in session:
        return {'success': False, 'message': 'Not logged in'}, 401
    
    data = request.get_json()
    report_id = data.get('report_id')
    user_id = session.get('user_id')
    
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    
    # Check if user is verified
    cursor.execute("SELECT account_status FROM accounts WHERE account_id = %s", (user_id,))
    user = cursor.fetchone()
    if not user or user.get('account_status') != 'Verified':
        con.close()
        return {'success': False, 'message': 'Account pending verification'}, 403
        
    
    # Check if already upvoted
    cursor.execute("SELECT upvote_id FROM upvotes WHERE report_id=%s AND account_id=%s", (report_id, user_id))
    existing = cursor.fetchone()
    
    if existing:
        cursor.execute("DELETE FROM upvotes WHERE upvote_id=%s", (existing['upvote_id'],))
        action = 'removed'
    else:
        cursor.execute("INSERT INTO upvotes (report_id, account_id) VALUES (%s, %s)", (report_id, user_id))
        action = 'added'
        
    con.commit()
    
    cursor.execute("SELECT COUNT(*) as count FROM upvotes WHERE report_id=%s", (report_id,))
    new_count = cursor.fetchone()['count']
    con.close()
    
    return {'success': True, 'action': action, 'count': new_count}

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        con = connect_db()
        cursor = con.cursor()
        data = request.form

        email = data.get('e', '').strip()
        phone = data.get('pn', '').strip()
        password = data.get('pass', '')
        conpass = data.get('conpass', '')

        # Must provide at least email or phone
        if not email and not phone:
            return render_template(
                'register.html',
                error="Enter email or phone number"
            )

        # Check existing credentials FIRST
        dupq = """
            SELECT emorph
            FROM usercreds
            WHERE emorph = %s OR emorph = %s
        """

        cursor.execute(dupq, (email, phone))
        duplicate = cursor.fetchone()

        if duplicate:
            return render_template(
                'register.html',
                error="The email or phone number has been already registered"
            )

        # Password validation
        if len(password) < 8:
            return render_template(
                'register.html',
                error="Password must be at least 8 characters long"
            )

        if password != conpass:
            return render_template(
                'register.html',
                error="Passwords do not match"
            )

        # Hash password
        hashed_pw = hash_password_bcrypt(password)

        # Insert temporary registration
        query = """
            INSERT INTO registration (email_address, phone_number, password)
            VALUES (%s, %s, %s)
        """

        values = (
            email if email else None,
            phone if phone else None,
            hashed_pw
        )

        cursor.execute(query, values)
        con.commit()

        return redirect(url_for('login'))

    return render_template('register.html')

@app.route('/complete_details', methods=['GET', 'POST'])
def cd():
    con = connect_db()
    cursor = con.cursor(buffered=True)
    reg_id = request.args.get('reg_id') or request.form.get('reg_id')
    
    if not reg_id:
        return redirect(url_for('login'))
    
    # Fetch reg_user data for display
    cred = "SELECT email_address, phone_number FROM registration WHERE registration_id = %s"
    cursor.execute(cred, (reg_id,))
    reg_user = cursor.fetchone()
    
    if request.method == 'GET':
        if reg_user:
            u = reg_user[0] or reg_user[1]
            return render_template('complete_details.html', u=u, reg_id=reg_id, email=reg_user[0] or '', phone=reg_user[1] or '')
        return redirect(url_for('login'))
    
    else:  # POST
        data = request.form
        
        # Handle cropped profile photo base64
        cropped_image = data.get('cropped_image')
        pp_path = None
        if cropped_image and cropped_image.startswith('data:image'):
            import base64
            header, encoded = cropped_image.split(",", 1)
            file_ext = header.split(";")[0].split("/")[1]
            filename = f"profile_{reg_id}_{int(datetime.now().timestamp())}.{file_ext}"
            os.makedirs('static/uploads', exist_ok=True)
            pp_path = f"static/uploads/{filename}"
            with open(pp_path, "wb") as fh:
                fh.write(base64.b64decode(encoded))
              
        # Insert into accounts with 'Incomplete' status since ID is not yet uploaded
        cursor.execute("""
            INSERT INTO accounts (account_id, first_name, last_name, middle_name, dob, profile_photo, house_no, street_name, province, city, barangay, account_status) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            reg_id, data.get('fname'), data.get('lname'), data.get('mname'), data.get('dob'), pp_path,
            data.get('house_no'), data.get('street_name'), data.get('province_name'), data.get('city_name'), data.get('barangay_name'),
            'Incomplete'
        ))
        
        cursor.execute("SELECT account_id FROM accounts WHERE account_id = %s", (reg_id,))
        newid = cursor.fetchone()[0]
        
        # Add contacts for email and phone
        cursor.execute(
            "INSERT INTO contacts (account_id, email, phone_number) VALUES (%s, %s, %s)", 
            (newid, reg_user[0] if reg_user[0] else None, reg_user[1] if reg_user[1] else None)
        )
                
        cursor.execute("""
            INSERT INTO password (password) 
            SELECT password FROM registration WHERE registration_id = %s
        """, (reg_id,))
        
        cursor.execute("SELECT MAX(pass_id) FROM password")
        new_pass_id = cursor.fetchone()[0]
        
        # Ensure credentials and user_role exist for ALL usercreds related to this account
        cursor.execute("""
            SELECT u.user_id FROM usercreds u
            JOIN contacts c ON u.contact_id = c.contact_id
            WHERE c.account_id = %s
        """, (newid,))
        user_ids = [row[0] for row in cursor.fetchall()]
        
        for uid in user_ids:
            cursor.execute("SELECT cred_id, pass_id FROM credentials WHERE user_id = %s", (uid,))
            row = cursor.fetchone()
            if not row:
                cursor.execute("INSERT INTO credentials (user_id, pass_id, is_locked) VALUES (%s, %s, 'N')", (uid, new_pass_id))
            elif row[1] is None:
                cursor.execute("UPDATE credentials SET pass_id = %s WHERE cred_id = %s", (new_pass_id, row[0]))
            
            cursor.execute("SELECT * FROM user_role WHERE user_id = %s", (uid,))
            if not cursor.fetchone():
                cursor.execute("INSERT INTO user_role (user_id, role_id) VALUES (%s, %s)", (uid, 1))
        
        # Finally delete registration record
        cursor.execute("DELETE FROM registration WHERE registration_id = %s", (reg_id,))
        con.commit()
        con.close()
        
        session['pending_user_id'] = newid # Pass to next step
        return redirect(url_for('id_verification'))

@app.route('/id_verification', methods=['GET', 'POST'])
def id_verification():
    if request.method == 'GET':
        user_id = session.get('pending_user_id')
        if not user_id:
            return redirect(url_for('login'))
        return render_template('id_verification.html', user_id=user_id)
    
    else: # POST
        user_id = request.form.get('user_id')
        if not user_id:
            return redirect(url_for('login'))
            
        id_proof = request.files.get('id_proof')
        id_path = None
        if id_proof and id_proof.filename:
            os.makedirs('static/uploads/id_proofs', exist_ok=True)
            id_path = f"static/uploads/id_proofs/{id_proof.filename}"
            id_proof.save(id_path)
            
            con = connect_db()
            cursor = con.cursor(buffered=True)
            cursor.execute("UPDATE accounts SET id_proof_path = %s, account_status = 'Pending' WHERE account_id = %s", (id_path, user_id))
            con.commit()
            
            # Auto login the user
            cursor.execute("SELECT first_name FROM accounts WHERE account_id = %s", (user_id,))
            user_name = cursor.fetchone()[0]
            con.close()
            
            session['user_id'] = user_id
            session['user_name'] = user_name
            session['user_role'] = 'resident'
            session.pop('pending_user_id', None)
            
            return redirect(url_for('dashboard'))
            
        return render_template('id_verification.html', user_id=user_id, error="Please upload a valid ID")

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/lgu_announcements', methods=['GET', 'POST'])
def lgu_announcements():
    if 'user_id' not in session or session.get('user_role') != 'lgu':
        return redirect(url_for('login'))
    
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    
    if request.method == 'POST':
        category = request.form.get('category', '').strip()
        title = request.form.get('title', '').strip()
        content = request.form.get('content', '').strip()
        account_id = session.get('user_id')
        
        if title and content:
            cursor.execute(
                "INSERT INTO announcements (account_id, category, ann_title, ann_desc, submitted_at) VALUES (%s, %s, %s, %s, %s)",
                (account_id, category, title, content, datetime.now())
            )
            con.commit()
            
        return redirect(url_for('lgu_announcements'))
        
    cursor.execute("SELECT first_name, last_name, profile_photo FROM accounts WHERE account_id = %s", (session['user_id'],))
    user_rows = cursor.fetchall()
    user_data = user_rows[0] if user_rows else None
    
    cursor.execute("SELECT * FROM announcements ORDER BY submitted_at DESC")
    announcements = cursor.fetchall()
    for ann in announcements:
        if ann['submitted_at']:
            ann['formatted_date'] = ann['submitted_at'].strftime('%B %d, %Y')
        else:
            ann['formatted_date'] = ''
            
    con.close()
    
    return render_template('lgu_announcements.html', user=user_data, announcements=announcements)

@app.route('/delete_announcement/<int:ann_id>', methods=['POST'])
def delete_announcement(ann_id):
    if 'user_id' not in session or session.get('user_role') != 'lgu':
        return redirect(url_for('login'))
        
    con = connect_db()
    cursor = con.cursor()
    cursor.execute("DELETE FROM announcements WHERE ann_id = %s", (ann_id,))
    con.commit()
    con.close()
    return redirect(url_for('lgu_announcements'))

@app.route('/barangay_map')
def barangay_map():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    con = connect_db()
    cursor = con.cursor(dictionary=True)
    cursor.execute(
        "SELECT first_name, last_name, profile_photo FROM accounts WHERE account_id = %s",
        (session['user_id'],)
    )
    user_rows = cursor.fetchall()
    user_data = user_rows[0] if user_rows else None

    # For leaflet pins: return rich report details + locname + x/y (location coordinates)
    rep = """
        SELECT 
            r.report_id, 
            r.title, 
            r.category, 
            r.description, 
            r.image_url, 
            r.status, 
            DATE_FORMAT(r.created_at, '%b %d, %Y • %h:%i %p') AS created_at_str,
            l.locname, 
            l.x, 
            l.y, 
            CONCAT(a.first_name, ' ', a.last_name) AS reporter_name,
            a.profile_photo AS reporter_photo
        FROM reports r
        JOIN location l ON r.location_id = l.location_id
        LEFT JOIN accounts a ON r.account_id = a.account_id
        WHERE l.x IS NOT NULL AND l.y IS NOT NULL
        ORDER BY r.created_at DESC
    """
    cursor.execute(rep)
    rloc = cursor.fetchall()

    # Sanitize any bytes/None for JSON serialization
    for r in rloc:
        for k, v in list(r.items()):
            if isinstance(v, (bytes, bytearray)):
                r[k] = v.decode('utf-8', errors='ignore') if v else ""
            elif v is None:
                r[k] = ""

    # DEBUG: verify DB data for leaflet pins
    print('[/barangay_map] rloc rows:', len(rloc))
    if rloc:
        print('[/barangay_map] first rloc row:', rloc[0])

    con.close()

    return render_template('barangay_map.html', user=user_data, rloc=rloc)


@app.route('/lgu_verifications')
def lgu_verifications():
    if 'user_id' not in session or session.get('user_role') != 'lgu':
        return redirect(url_for('login'))
        
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    
    # Get LGU user info
    cursor.execute("SELECT first_name, last_name, profile_photo FROM accounts WHERE account_id = %s", (session['user_id'],))
    user_rows = cursor.fetchall()
    user_data = user_rows[0] if user_rows else None
    
    # Get pending users
    cursor.execute("""
        SELECT a.account_id, a.first_name, a.last_name, a.profile_photo, a.id_proof_path, c.email, c.phone_number 
        FROM accounts a
        LEFT JOIN contacts c ON a.account_id = c.account_id
        WHERE a.account_status = 'Pending'
    """)
    pending_users = cursor.fetchall()
    con.close()
    
    return render_template('lgu_verifications.html', user=user_data, pending_users=pending_users)

@app.route('/verify_user', methods=['POST'])
def verify_user():
    if 'user_id' not in session or session.get('user_role') != 'lgu':
        return redirect(url_for('login'))
        
    account_id = request.form.get('account_id')
    action = request.form.get('action') # 'accept' or 'reject'
    
    if account_id and action:
        con = connect_db()
        cursor = con.cursor()
        new_status = 'Verified' if action == 'accept' else 'Rejected'
        cursor.execute("UPDATE accounts SET account_status = %s WHERE account_id = %s", (new_status, account_id))
        con.commit()
        con.close()
        
    return redirect(url_for('lgu_verifications'))

@app.route('/lgu_analytics')
def lgu_analytics():
    if 'user_id' not in session or session.get('user_role') != 'lgu':
        return redirect(url_for('login'))

    con = connect_db()
    cursor = con.cursor(dictionary=True)
    account_id = session.get('user_id')

    # --- User profile ---
    cursor.execute(
        "SELECT first_name, last_name, profile_photo FROM accounts WHERE account_id = %s",
        (account_id,)
    )
    user_rows = cursor.fetchall()
    user_data = user_rows[0] if user_rows else None

    # --- Overall counts ---
    cursor.execute("SELECT COUNT(*) AS total FROM reports")
    total = cursor.fetchone()['total'] or 0

    cursor.execute("SELECT COUNT(*) AS cnt FROM reports WHERE LOWER(status) = 'pending'")
    pending = cursor.fetchone()['cnt'] or 0

    cursor.execute("SELECT COUNT(*) AS cnt FROM reports WHERE LOWER(status) IN ('in progress','inprogress','in_progress')")
    in_progress = cursor.fetchone()['cnt'] or 0

    cursor.execute("SELECT COUNT(*) AS cnt FROM reports WHERE LOWER(status) = 'resolved'")
    resolved = cursor.fetchone()['cnt'] or 0

    cursor.execute("SELECT COUNT(*) AS cnt FROM reports WHERE LOWER(status) IN ('denied', 'rejected', 'declined', 'cancelled', 'spam', 'invalid')")
    denied = cursor.fetchone()['cnt'] or 0

    cursor.execute("SELECT COUNT(*) AS cnt FROM reports WHERE DATE(created_at) = CURDATE()")
    today_count = cursor.fetchone()['cnt'] or 0

    cursor.execute("SELECT COUNT(*) AS cnt FROM reports WHERE YEARWEEK(created_at,1)=YEARWEEK(CURDATE(),1)")
    this_week = cursor.fetchone()['cnt'] or 0

    # This month / last month
    cursor.execute("""
        SELECT COUNT(*) AS cnt FROM reports
        WHERE YEAR(created_at) = YEAR(CURDATE()) AND MONTH(created_at) = MONTH(CURDATE())
    """)
    this_month = cursor.fetchone()['cnt'] or 0

    cursor.execute("""
        SELECT COUNT(*) AS cnt FROM reports
        WHERE YEAR(created_at) = YEAR(DATE_SUB(CURDATE(), INTERVAL 1 MONTH))
          AND MONTH(created_at) = MONTH(DATE_SUB(CURDATE(), INTERVAL 1 MONTH))
    """)
    last_month = cursor.fetchone()['cnt'] or 0

    # --- Engagement ---
    cursor.execute("SELECT COUNT(*) AS cnt FROM upvotes")
    total_upvotes = cursor.fetchone()['cnt'] or 0

    cursor.execute("SELECT COUNT(*) AS cnt FROM feedbacks")
    total_comments = cursor.fetchone()['cnt'] or 0

    cursor.execute("SELECT COUNT(DISTINCT account_id) AS cnt FROM reports")
    total_residents = cursor.fetchone()['cnt'] or 0

    # --- Monthly trend (last 6 months) ---
    cursor.execute("""
        SELECT DATE_FORMAT(created_at, '%b %Y') AS month_label,
               YEAR(created_at) AS yr, MONTH(created_at) AS mo,
               COUNT(*) AS cnt,
               SUM(CASE WHEN LOWER(status)='resolved' THEN 1 ELSE 0 END) AS resolved_cnt
        FROM reports
        WHERE created_at >= DATE_SUB(CURDATE(), INTERVAL 6 MONTH)
        GROUP BY yr, mo, month_label ORDER BY yr ASC, mo ASC
    """)
    trend_rows = cursor.fetchall()
    monthly_labels   = [r['month_label']   for r in trend_rows]
    monthly_counts   = [r['cnt']           for r in trend_rows]
    monthly_resolved = [r['resolved_cnt']  for r in trend_rows]

    # --- Day-of-week distribution ---
    cursor.execute("""
        SELECT DAYOFWEEK(created_at) AS dow, COUNT(*) AS cnt
        FROM reports GROUP BY dow ORDER BY dow
    """)
    dow_map = {r['dow']: r['cnt'] for r in cursor.fetchall()}
    dow_labels = ['Sun','Mon','Tue','Wed','Thu','Fri','Sat']
    dow_counts = [dow_map.get(i,0) for i in range(1,8)]

    # --- Category breakdown ---
    cat_colors = {
        'road':       '#fbbf24',
        'sanitation': '#34d399',
        'safety':     '#f87171',
        'flooding':   '#60a5fa',
        'lighting':   '#a78bfa',
        'other':      '#94a3b8',
    }
    cursor.execute("""
        SELECT LOWER(REPLACE(category,' ','')) AS cat_key, category, COUNT(*) AS cnt,
               SUM(CASE WHEN LOWER(status)='resolved' THEN 1 ELSE 0 END) AS res_cnt
        FROM reports GROUP BY cat_key, category ORDER BY cnt DESC
    """)
    cat_rows = cursor.fetchall()
    category_data = [
        {'name': r['category'], 'count': r['cnt'], 'resolved': r['res_cnt'],
         'color': cat_colors.get(r['cat_key'], '#8b0000')}
        for r in cat_rows
    ]
    category_names  = [c['name']  for c in category_data]
    category_counts = [c['count'] for c in category_data]

    # --- Top locations ---
    cursor.execute("""
        SELECT l.locname, COUNT(*) AS cnt
        FROM reports r
        JOIN location l ON r.location_id = l.location_id
        WHERE l.locname IS NOT NULL AND l.locname != ''
        GROUP BY l.locname ORDER BY cnt DESC LIMIT 7
    """)
    top_locations = [
        {'locname': r['locname'], 'count': r['cnt']}
        for r in cursor.fetchall()
    ]

    # --- Top reporters ---
    cursor.execute("""
        SELECT a.first_name, a.last_name, a.profile_photo, COUNT(*) AS cnt
        FROM reports r JOIN accounts a ON r.account_id=a.account_id
        GROUP BY r.account_id,a.first_name,a.last_name,a.profile_photo
        ORDER BY cnt DESC LIMIT 5
    """)
    top_reporters = [
        {'name': f"{r['first_name']} {r['last_name']}",
         'count': r['cnt'],
         'initial': (r['first_name'] or 'U')[0].upper(),
         'photo': r['profile_photo']}
        for r in cursor.fetchall()
    ]

    # --- Oldest pending ---
    cursor.execute("""
        SELECT r.title, r.category, r.created_at, a.first_name, a.last_name,
               DATEDIFF(CURDATE(), DATE(r.created_at)) AS days_old
        FROM reports r JOIN accounts a ON r.account_id=a.account_id
        WHERE LOWER(r.status)='pending' ORDER BY r.created_at ASC LIMIT 5
    """)
    oldest_pending = cursor.fetchall()

    # --- Recent 10 reports ---
    cursor.execute("""
        SELECT r.report_id, r.title, r.category, r.status, r.created_at,
               a.first_name, a.last_name, l.locname
        FROM reports r
        JOIN accounts a ON r.account_id = a.account_id
        LEFT JOIN location l ON r.location_id = l.location_id
        ORDER BY r.created_at DESC LIMIT 10
    """)
    recent_reports = cursor.fetchall()

    con.close()

    stats = {
        'total': total, 'pending': pending,
        'in_progress': in_progress, 'resolved': resolved,
        'denied': denied,
        'today': today_count, 'this_week': this_week,
        'this_month': this_month, 'last_month': last_month,
        'total_upvotes': total_upvotes,
        'total_comments': total_comments,
        'total_residents': total_residents,
    }

    return render_template(
        'lgu_analytics.html',
        user=user_data,
        stats=stats,
        monthly_labels=monthly_labels,
        monthly_counts=monthly_counts,
        monthly_resolved=monthly_resolved,
        dow_labels=dow_labels,
        dow_counts=dow_counts,
        category_data=category_data,
        category_names=category_names,
        category_counts=category_counts,
        top_locations=top_locations,
        top_reporters=top_reporters,
        oldest_pending=oldest_pending,
        recent_reports=recent_reports,
    )


@app.route('/lgu_map')
def lgu_map():
    if 'user_id' not in session or session.get('user_role') != 'lgu':
        return redirect(url_for('login'))

    con = connect_db()
    cursor = con.cursor(dictionary=True)

    cursor.execute(
        "SELECT first_name, last_name, profile_photo FROM accounts WHERE account_id = %s",
        (session['user_id'],)
    )
    user_rows = cursor.fetchall()
    user_data = user_rows[0] if user_rows else None

    # For leaflet pins: return rich report details + locname + x/y (location coordinates)
    rep = """
        SELECT 
            r.report_id, 
            r.title, 
            r.category, 
            r.description, 
            r.image_url, 
            r.status, 
            DATE_FORMAT(r.created_at, '%b %d, %Y • %h:%i %p') AS created_at_str,
            l.locname, 
            l.x, 
            l.y, 
            CONCAT(a.first_name, ' ', a.last_name) AS reporter_name,
            a.profile_photo AS reporter_photo
        FROM reports r
        JOIN location l ON r.location_id = l.location_id
        LEFT JOIN accounts a ON r.account_id = a.account_id
        WHERE l.x IS NOT NULL AND l.y IS NOT NULL
        ORDER BY r.created_at DESC
    """
    cursor.execute(rep)
    rloc = cursor.fetchall()

    # Sanitize any bytes/None for JSON serialization
    for r in rloc:
        for k, v in list(r.items()):
            if isinstance(v, (bytes, bytearray)):
                r[k] = v.decode('utf-8', errors='ignore') if v else ""
            elif v is None:
                r[k] = ""

    con.close()

    return render_template('lgu_map.html', user=user_data, rloc=rloc)




@app.route('/announcements')
def announcements():
    if 'user_id' not in session:
        return redirect(url_for('login'))
        
    con = connect_db()
    cursor = con.cursor(dictionary=True)
    cursor.execute("SELECT first_name, last_name, profile_photo FROM accounts WHERE account_id = %s", (session['user_id'],))
    user_rows = cursor.fetchall()
    user_data = user_rows[0] if user_rows else None
    
    cursor.execute("""
        SELECT a.*, acc.first_name, acc.last_name 
        FROM announcements a
        LEFT JOIN accounts acc ON a.account_id = acc.account_id
        ORDER BY a.submitted_at DESC
    """)
    db_anns = cursor.fetchall()
    
    # Format for JSON
    anns_list = []
    emoji_map = {
        'General': '📢',
        'Health': '🏥',
        'Infrastructure': '🛣️',
        'Events': '🎉',
        'Safety': '🚨'
    }
    
    for row in db_anns:
        cat = row['category'] or 'General'
        emoji = emoji_map.get(cat, '📢')
        author = f"{row['first_name']} {row['last_name']}" if row['first_name'] else "Barangay Picaleon"
        date_str = row['submitted_at'].strftime('%B %d, %Y') if row['submitted_at'] else ''
        
        # Build JSON item
        anns_list.append({
            'cat': cat.lower(),
            'emoji': emoji,
            'tag': f"{emoji} {cat}",
            'title': row['ann_title'],
            'excerpt': (row['ann_desc'][:100] + '...') if row['ann_desc'] and len(row['ann_desc']) > 100 else (row['ann_desc'] or ''),
            'date': date_str,
            'author': author,
            'views': '-',
            'content': f"<p>{row['ann_desc']}</p>"
        })
        
    con.close()
    
    import json
    anns_json = json.dumps(anns_list)
    
    return render_template('announcements.html', user=user_data, announcements_json=anns_json)

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000, debug=True)

